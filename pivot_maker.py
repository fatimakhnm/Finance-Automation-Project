import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell

# Load the Excel file
file_path = r'C:\Users\MK\Desktop\Python\PY4E\Projects\ee_data.xlsx'
df = pd.read_excel(file_path)

# Remove duplicate columns if any
df = df.loc[:, ~df.columns.duplicated()]

# Remove duplicate rows based on all columns
df = df.drop_duplicates()

# Convert currency columns to numeric, removing currency symbols and commas
currency_columns = ['Position Budget', 'YTD Actuals', 'Actual for Month', 'EE Annual Salary']
for col in currency_columns:
    df[col] = df[col].replace(r'[^\d.]', '', regex=True).astype(float)

# Calculate additional KPIs
df['Remainder Forecasted'] = (df['YTD Actuals'] / 6) * 12
df['Budget Minus Remainder Forecasted'] = df['Position Budget'] - df['Remainder Forecasted']
df['Forecasted Saving/Underspend'] = df['Position Budget'] - (df['YTD Actuals'] + df['Remainder Forecasted'])

# Create a pivot table with the updated calculations
pivot_table = pd.pivot_table(
    df,
    index=['Department', 'Job Title'],
    values=[
        'Position Budget',
        'YTD Actuals',
        'Remainder Forecasted'
    ],
    aggfunc={
        'Position Budget': 'sum',
        'YTD Actuals': 'sum',
        'Remainder Forecasted': 'sum'
    },
    margins=True,
    margins_name='Subtotal'
)

# Calculate Forecasted Saving/Underspend separately and add it to the pivot table
pivot_table['Forecasted Saving/Underspend'] = pivot_table['Position Budget'] - pivot_table['Remainder Forecasted']

# Ensure all columns are formatted correctly
pivot_table = pivot_table.round({
    'Position Budget': 0,
    'YTD Actuals': 0,
    'Remainder Forecasted': 0,
    'Forecasted Saving/Underspend': 0
})

# Format the pivot table
def format_number(x):
    return "{:,.0f}".format(x)

for col in pivot_table.columns:
    pivot_table[col] = pivot_table[col].apply(format_number)

# Save the updated pivot table
pivot_output_path = r'C:\Users\MK\Desktop\Python\PY4E\Projects\updated_ee_pivot_table.xlsx'
with pd.ExcelWriter(pivot_output_path, engine='openpyxl') as writer:
    pivot_table.to_excel(writer, sheet_name='UpdatedPivotTable', index=True)

# Load the workbook and select the sheet
workbook = load_workbook(pivot_output_path)
sheet = workbook['UpdatedPivotTable']

# Define fill colors for conditional formatting
pastel_green_fill = PatternFill(start_color='B2FFB2', end_color='B2FFB2', fill_type='solid')  # Pastel green
pastel_red_fill = PatternFill(start_color='FFB2B2', end_color='FFB2B2', fill_type='solid')  # Pastel red

# Apply conditional formatting to the 'Forecasted Saving/Underspend' column (Column F)
column_letter = 'F'  # The column where 'Forecasted Saving/Underspend' is located
column_index = 6  # Column F is the 6th column in Excel

for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, max_row=sheet.max_row):
    for cell in row:
        if not isinstance(cell, MergedCell):  # Avoid merged cells
            try:
                value = float(cell.value.replace(',', ''))
                if value < 0:  # Indicates overspend
                    cell.fill = pastel_red_fill
                else:  # Indicates saving
                    cell.fill = pastel_green_fill
            except (ValueError, AttributeError):
                # Skip cells that cannot be converted to float
                continue

# Save the workbook with formatting
workbook.save(pivot_output_path)

print("Pivot table with KPIs, formatting, and conditional formatting has been created and saved successfully.")
